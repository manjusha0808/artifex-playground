"""
Validate the crawler output Excel file for quality issues.
"""
import sys
sys.path.insert(0, '..')

from openpyxl import load_workbook
from collections import Counter

def validate_excel(filepath):
    """Validate Excel file for duplicates, missing data, and quality issues."""
    
    print(f"📊 Validating: {filepath}\n")
    
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        total_rows = ws.max_row - 1  # Exclude header
        print(f"✅ Total data rows: {total_rows}")
        
        # Get column headers
        headers = [cell.value for cell in ws[1]]
        print(f"📋 Columns: {headers}\n")
        
        # Collect data
        xpaths = []
        pages = []
        hierarchies = []
        element_ids = []
        element_names = []
        missing_xpath = 0
        missing_hierarchy = 0
        missing_page = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
                
            hierarchy = row[0]
            page = row[1]
            element_name = row[2]
            xpath = row[3]
            elem_id = row[4]
            
            hierarchies.append(hierarchy)
            pages.append(page)
            element_names.append(element_name)
            
            if xpath:
                xpaths.append(xpath)
            else:
                missing_xpath += 1
            
            if not hierarchy or hierarchy == 'Unknown':
                missing_hierarchy += 1
            
            if not page:
                missing_page += 1
            
            if elem_id:
                element_ids.append(elem_id)
        
        print(f"🔍 Data Quality Check:")
        print(f"   - Total XPaths extracted: {len(xpaths)}")
        print(f"   - Missing XPaths: {missing_xpath}")
        print(f"   - Missing hierarchy: {missing_hierarchy}")
        print(f"   - Missing page names: {missing_page}")
        print(f"   - Unique pages: {len(set(pages))}")
        print(f"   - Elements with IDs: {len(element_ids)}\n")
        
        # Check for duplicate XPaths
        xpath_counts = Counter(xpaths)
        duplicates = {xpath: count for xpath, count in xpath_counts.items() if count > 1}
        
        if duplicates:
            print(f"⚠️ DUPLICATE XPATHS FOUND: {len(duplicates)} XPaths appear multiple times")
            print(f"   Top 10 duplicates:")
            for xpath, count in sorted(duplicates.items(), key=lambda x: x[1], reverse=True)[:10]:
                print(f"   - {count}x: {xpath[:80]}")
        else:
            print(f"✅ No duplicate XPaths found!")
        
        print(f"\n📄 Page-level duplicates check:")
        # Group by page and check for duplicates within each page
        page_xpath_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            page = row[1]
            xpath = row[3]
            if page and xpath:
                if page not in page_xpath_map:
                    page_xpath_map[page] = []
                page_xpath_map[page].append(xpath)
        
        pages_with_dupes = 0
        for page, page_xpaths in page_xpath_map.items():
            xpath_counts = Counter(page_xpaths)
            page_dupes = {xpath: count for xpath, count in xpath_counts.items() if count > 1}
            if page_dupes:
                pages_with_dupes += 1
                if pages_with_dupes <= 5:  # Show first 5
                    print(f"   ⚠️ Page '{page}' has duplicates:")
                    for xpath, count in list(page_dupes.items())[:3]:
                        print(f"      - {count}x: {xpath[:60]}")
        
        if pages_with_dupes == 0:
            print(f"   ✅ No page-level duplicates found!")
        else:
            print(f"   ⚠️ Total pages with duplicates: {pages_with_dupes}")
        
        # Sample data check
        print(f"\n📋 Sample data (first 5 rows):")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
            print(f"   Row {i}:")
            print(f"      Page: {row[1]}")
            print(f"      Element: {row[2]}")
            print(f"      XPath: {row[3][:80] if row[3] else 'MISSING'}")
        
        # Check for hallucinated data patterns
        print(f"\n🔍 Hallucination check:")
        suspicious_count = 0
        for xpath in xpaths[:100]:  # Check first 100
            if xpath and ('null' in xpath.lower() or 'undefined' in xpath.lower() or '//' in xpath[:3]):
                suspicious_count += 1
        
        if suspicious_count > 0:
            print(f"   ⚠️ Found {suspicious_count} potentially suspicious XPaths in sample")
        else:
            print(f"   ✅ No obvious hallucinated data detected in sample")
        
        print(f"\n{'='*60}")
        print(f"SUMMARY:")
        print(f"{'='*60}")
        print(f"✅ Total records: {total_rows}")
        print(f"{'✅' if missing_xpath == 0 else '⚠️'} Missing XPaths: {missing_xpath}")
        print(f"{'✅' if len(duplicates) == 0 else '⚠️'} Duplicate XPaths: {len(duplicates)}")
        print(f"{'✅' if pages_with_dupes == 0 else '⚠️'} Pages with duplicates: {pages_with_dupes}")
        print(f"{'='*60}")
        
    except FileNotFoundError:
        print(f"❌ File not found: {filepath}")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    import sys
    filepath = sys.argv[1] if len(sys.argv) > 1 else 'uiMap_selenium_fullrun_auto3.xlsx'
    validate_excel(filepath)
