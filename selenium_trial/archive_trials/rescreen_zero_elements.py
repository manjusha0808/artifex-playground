"""
Re-screen menu items that had 0 elements extracted during initial crawl.
Compares checkpoint (processed items) vs Excel (items with data) to find misses.
"""
import json
import openpyxl
import os

# Configuration
CHECKPOINT_FILE = "crawler_checkpoint.json"
EXCEL_FILE = "uiMap_selenium_fullrun_auto4_cleaned.xlsx"
OUTPUT_FILE = "items_to_rescreen.json"

def load_checkpoint():
    """Load the checkpoint file to get list of processed items."""
    if not os.path.exists(CHECKPOINT_FILE):
        print(f"❌ Checkpoint file not found: {CHECKPOINT_FILE}")
        return None, None
    
    with open(CHECKPOINT_FILE, 'r') as f:
        data = json.load(f)
    
    processed_items = set(data.get('processed_items', []))
    last_index = data.get('last_index', 0)
    
    print(f"📋 Checkpoint loaded:")
    print(f"   - Total processed: {len(processed_items)}")
    print(f"   - Last index: {last_index}")
    
    return processed_items, last_index

def load_excel_pages():
    """Load Excel file and get unique page names."""
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Excel file not found: {EXCEL_FILE}")
        return set()
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    pages_with_data = set()
    for i in range(2, ws.max_row + 1):
        page_name = ws[i][0].value
        if page_name:
            pages_with_data.add(page_name)
    
    print(f"📊 Excel analysis:")
    print(f"   - Total rows: {ws.max_row - 1}")
    print(f"   - Unique pages: {len(pages_with_data)}")
    
    return pages_with_data

def find_items_to_rescreen():
    """Compare processed items vs pages with data to find items that had 0 elements."""
    processed_items, last_index = load_checkpoint()
    if processed_items is None:
        return
    
    pages_with_data = load_excel_pages()
    
    # Items that were processed but have no data = had 0 elements
    zero_element_items = processed_items - pages_with_data
    
    print(f"\n🔍 Analysis:")
    print(f"   - Processed but no data: {len(zero_element_items)}")
    print(f"   - Pages with data: {len(pages_with_data)}")
    
    if zero_element_items:
        print(f"\n📝 Items to rescreen ({len(zero_element_items)}):")
        
        # Save to file
        output_data = {
            "items_to_rescreen": sorted(list(zero_element_items)),
            "count": len(zero_element_items),
            "reason": "These items were processed but yielded 0 elements"
        }
        
        with open(OUTPUT_FILE, 'w') as f:
            json.dump(output_data, f, indent=2)
        
        print(f"\n✅ Saved to: {OUTPUT_FILE}")
        
        # Show first 20 items
        print(f"\nFirst 20 items to rescreen:")
        for i, item in enumerate(sorted(list(zero_element_items))[:20], 1):
            print(f"   {i}. {item}")
        
        if len(zero_element_items) > 20:
            print(f"   ... and {len(zero_element_items) - 20} more")
    else:
        print("\n✅ All processed items have data in Excel!")

if __name__ == "__main__":
    print("=" * 70)
    print("RESCREENING ANALYSIS - Find items with 0 elements")
    print("=" * 70)
    print()
    
    find_items_to_rescreen()
    
    print("\n" + "=" * 70)
    print("Next step: Use 'rescreen_crawler.py' to re-process these items")
    print("=" * 70)
