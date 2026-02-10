"""Selenium-based T24 Menu Crawler - Single-threaded approach with checkpoint/resume."""

import time
import os
import json
from typing import List, Dict, Set
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# Load environment variables
load_dotenv()

# Checkpoint file
CHECKPOINT_FILE = 'crawler_checkpoint.json'


def load_checkpoint() -> Dict:
    """Load checkpoint data from file."""
    if os.path.exists(CHECKPOINT_FILE):
        try:
            with open(CHECKPOINT_FILE, 'r') as f:
                return json.load(f)
        except:
            return {'processed_items': [], 'last_index': 0}
    return {'processed_items': [], 'last_index': 0}


def save_checkpoint(processed_items: List[str], last_index: int):
    """Save checkpoint data to file with immediate flush."""
    try:
        with open(CHECKPOINT_FILE, 'w') as f:
            # Remove indent for faster writing, remove fsync for speed
            json.dump({
                'processed_items': processed_items,
                'last_index': last_index,
                'timestamp': time.strftime('%Y-%m-%d %H:%M:%S')
            }, f)
            f.flush()  # Flush to OS buffer (fast)
            # Removed os.fsync() - too slow, flush() is enough
    except Exception as e:
        print(f"⚠️ Failed to save checkpoint: {e}")


def clear_checkpoint():
    """Clear checkpoint file after successful completion."""
    try:
        if os.path.exists(CHECKPOINT_FILE):
            os.remove(CHECKPOINT_FILE)
            print("✅ Checkpoint cleared")
    except Exception as e:
        print(f"⚠️ Failed to clear checkpoint: {e}")


def load_existing_data(filepath: str) -> List[Dict]:
    """Load existing data from Excel file if resuming."""
    if os.path.exists(filepath):
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            data = []
            
            # Skip header row - new structure without hierarchy
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Check if row has data
                    data.append({
                        'page': row[0],
                        'elementName': row[1],
                        'relativeXpath': row[2],
                        'id': row[3],
                        'name': row[4],
                        'className': row[5],
                        'tagName': row[6],
                        'type': row[7]
                    })
            
            print(f"📂 Loaded {len(data)} existing records from {filepath}")
            return data
        except Exception as e:
            print(f"⚠️ Failed to load existing data: {e}")
            return []
    return []


def setup_driver():
    """Setup Chrome driver with options."""
    chrome_options = Options()
    # chrome_options.add_argument('--headless')  # Comment out to see the browser
    chrome_options.add_argument('--start-maximized')
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    
    driver = webdriver.Chrome(options=chrome_options)
    driver.implicitly_wait(3)
    return driver


def login(driver, url: str, username: str, password: str):
    """Login to T24 application."""
    print(f"🌐 Navigating to {url}")
    driver.get(url)
    
    time.sleep(2)
    
    print(f"🔐 Logging in as {username}")
    
    # Find and fill login fields
    user_field = driver.find_element(By.NAME, 'signOnName')
    pass_field = driver.find_element(By.NAME, 'password')
    
    user_field.send_keys(username)
    pass_field.send_keys(password)
    
    # Submit
    submit_btn = driver.find_element(By.CSS_SELECTOR, 'input[type="submit"]')
    submit_btn.click()
    
    time.sleep(3)
    print("✅ Login successful")


def get_menu_frame(driver, wait=False):
    """Switch to menu frame."""
    # Only wait on first call or when explicitly requested
    if wait:
        time.sleep(2)
    
    # Find menu frame
    frames = driver.find_elements(By.TAG_NAME, 'frame')
    if not frames:
        return False
    
    # Debug on first call only (when wait=True)
    if wait:
        print(f"📍 Found {len(frames)} frames")
        for i, frame in enumerate(frames):
            frame_name = frame.get_attribute('name')
            print(f"  Frame {i}: {frame_name}")
    
    # Switch to menu frame
    for frame in frames:
        frame_name = frame.get_attribute('name')
        if 'menu' in frame_name.lower():
            if wait:
                print(f"✅ Switching to menu frame: {frame_name}")
            driver.switch_to.frame(frame)
            return True
    
    return False


def extract_xpaths_from_element(element) -> Dict:
    """Extract XPath and attributes from a single element."""
    data = {}
    
    try:
        # Get attributes
        data['id'] = element.get_attribute('id') or ''
        data['name'] = element.get_attribute('name') or ''
        data['className'] = element.get_attribute('class') or ''
        data['tagName'] = element.tag_name
        data['type'] = element.get_attribute('type') or ''
        
        # Generate element name with prefix
        elem_id = data['id']
        elem_name_attr = data['name']
        
        # Use ID or name attribute for naming, prefer ID
        identifier = elem_id or elem_name_attr or ''
        
        if data['tagName'] == 'input':
            input_type = data['type'].lower() or 'text'
            if input_type in ['text', 'password', 'email', 'number', 'tel', 'search', 'url']:
                prefix = 'txt'
            elif input_type == 'checkbox':
                prefix = 'chk'
            elif input_type == 'radio':
                prefix = 'rad'
            elif input_type in ['date', 'time', 'datetime-local']:
                prefix = 'dte'
            else:
                prefix = 'inp'
            data['elementName'] = f"{prefix}_{identifier}" if identifier else f"{prefix}_{input_type}"
        elif data['tagName'] == 'select':
            data['elementName'] = f"ddl_{identifier}" if identifier else 'ddl_select'
        elif data['tagName'] == 'textarea':
            data['elementName'] = f"txt_{identifier}" if identifier else 'txt_area'
        else:
            data['elementName'] = identifier or f"{data['tagName']}_field"
        
        # Generate XPath
        if data['id']:
            data['relativeXpath'] = f"//{data['tagName']}[@id='{data['id']}']"
        elif data['name']:
            data['relativeXpath'] = f"//{data['tagName']}[@name='{data['name']}']"
        else:
            data['relativeXpath'] = f"//{data['tagName']}[@class='{data['className']}']" if data['className'] else ''
        
    except Exception as e:
        print(f"    ⚠️ Error extracting from element: {e}")
    
    return data


def get_menu_hierarchy(li_element) -> str:
    """Extract the full menu hierarchy by traversing parent LI elements."""
    hierarchy = []
    current = li_element
    
    try:
        # Traverse up to collect all parent menu items
        while current:
            # Get text from current LI (could be from <a> or <span>)
            try:
                # Try to get text from <a> (leaf)
                link = current.find_element(By.XPATH, ".//a")
                text = link.text.strip()
                if text:
                    hierarchy.insert(0, text)
            except:
                # Try to get text from <span> (parent)
                try:
                    spans = current.find_elements(By.XPATH, ".//span")
                    for span in spans:
                        text = span.text.strip()
                        if text and 'ProcessMouseClick' not in (span.get_attribute('onclick') or ''):
                            if text not in hierarchy:
                                hierarchy.insert(0, text)
                            break
                except:
                    pass
            
            # Move to parent LI
            try:
                current = current.find_element(By.XPATH, "./parent::ul/parent::li")
            except:
                break
    except Exception as e:
        pass
    
    return ' > '.join(hierarchy) if hierarchy else 'Unknown'


def extract_xpaths_from_page(driver, page_name: str, global_seen_xpaths: set = None) -> List[Dict]:
    """Extract XPaths from input fields only (elements that accept user input)."""
    elements_data = []
    # Use global set if provided, otherwise create local one
    seen_rows = global_seen_xpaths if global_seen_xpaths is not None else set()
    
    # Single comprehensive selector to avoid querying same elements multiple times
    # Target only INPUT-accepting elements (no buttons, links, submit, etc.)
    selector = '''
        input[type="text"],
        input[type="password"],
        input[type="email"],
        input[type="number"],
        input[type="tel"],
        input[type="date"],
        input[type="time"],
        input[type="datetime-local"],
        input[type="search"],
        input[type="url"],
        input[type="checkbox"],
        input[type="radio"],
        input:not([type]),
        textarea,
        select
    '''
    
    try:
        elements = driver.find_elements(By.CSS_SELECTOR, selector)
        
        for elem in elements:
            data = extract_xpaths_from_element(elem)
            xpath = data.get('relativeXpath')
            
            if xpath:
                # Create unique key from ALL columns (page + xpath + element details)
                row_key = (
                    page_name,
                    xpath,
                    data.get('elementName', ''),
                    data.get('id', ''),
                    data.get('name', ''),
                    data.get('className', ''),
                    data.get('tagName', ''),
                    data.get('type', '')
                )
                
                # Only add if this exact row hasn't been seen
                if row_key not in seen_rows:
                    seen_rows.add(row_key)
                    data['page'] = page_name
                    elements_data.append(data)
            
    except Exception as e:
        print(f"    ⚠️ Error extracting elements: {e}")
    
    return elements_data


def export_to_excel(data: List[Dict], filepath: str, sheet_name: str = 'T24ModelBank'):
    """Export data to Excel - overwrites file with all data."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Headers
    headers = ['pageName', 'elementName', 'relativeXpath', 'elementId', 'elementNameAttr', 'className', 'tagName', 'inputType']
    ws.append(headers)
    
    # Data rows - deduplicate by complete row, not just XPath
    seen_rows = set()
    unique_count = 0
    for row in data:
        # Create key from all columns
        row_key = (
            row.get('page', ''),
            row.get('elementName', ''),
            row.get('relativeXpath', ''),
            row.get('id', ''),
            row.get('name', ''),
            row.get('className', ''),
            row.get('tagName', ''),
            row.get('type', '')
        )
        
        # Only write if this exact row hasn't been written
        if row_key not in seen_rows:
            seen_rows.add(row_key)
            ws.append([
                row.get('page', ''),
                row.get('elementName', ''),
                row.get('relativeXpath', ''),
                row.get('id', ''),
                row.get('name', ''),
                row.get('className', ''),
                row.get('tagName', ''),
                row.get('type', '')
            ])
            unique_count += 1
    
    wb.save(filepath)
    print(f"💾 Saved {unique_count} unique rows to {filepath}")


def export_stats_to_excel(stats_data: List[Dict], filepath: str):
    """Export page statistics to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Page Stats'
    
    # Headers
    ws.append(['pageName', 'xpath_count'])
    
    # Data rows
    for row in stats_data:
        ws.append([row['page'], row['count']])
    
    wb.save(filepath)


def expand_all_menus_recursive(driver):
    """Recursively expand ALL collapsed menu nodes by clicking expand icons."""
    print("🔧 Auto-expanding all menu items...")
    
    driver.switch_to.default_content()
    get_menu_frame(driver, wait=True)  # Only wait on initial call
    
    expanded_total = 0
    max_passes = 20  # Maximum expansion passes
    
    for pass_num in range(max_passes):
        # Find all span elements with ProcessMouseClick (these are expand/collapse controls)
        try:
            container = driver.find_element(By.XPATH, "//div[starts-with(@id,'pane')]")
            expand_icons = container.find_elements(By.XPATH, ".//span[contains(@onclick, 'ProcessMouseClick')]")
            
            if not expand_icons:
                print(f"✅ No more expandable items found")
                break
            
            expanded_this_pass = 0
            
            for icon in expand_icons:
                try:
                    # Check if the parent LI is collapsed (hidden children)
                    parent_li = icon.find_element(By.XPATH, "./ancestor::li[1]")
                    
                    # Check if there are hidden child UL elements (means it's collapsed)
                    child_uls = parent_li.find_elements(By.XPATH, "./ul")
                    
                    if child_uls:
                        # Check if child UL is visible
                        is_expanded = any(ul.is_displayed() for ul in child_uls)
                        
                        if not is_expanded:
                            # Click to expand
                            icon.click()
                            expanded_this_pass += 1
                            time.sleep(0.05)  # Small delay
                except:
                    continue
            
            expanded_total += expanded_this_pass
            print(f"   Pass {pass_num + 1}: Expanded {expanded_this_pass} nodes (Total: {expanded_total})")
            
            if expanded_this_pass == 0:
                print(f"✅ All menus fully expanded")
                break
            
            time.sleep(0.3)  # Wait for DOM updates
            
        except Exception as e:
            print(f"⚠️ Expansion error: {e}")
            break
    
    print(f"📂 Expansion complete: {expanded_total} total nodes expanded\n")
    return expanded_total


def crawl_menu():
    """Main crawler function with automatic recursive menu expansion."""
    # Configuration
    URL = os.getenv('APP_URL', 'http://10.0.251.41:18080/BrowserWeb/servlet/BrowserServlet')
    USERNAME = os.getenv('APP_USERNAME', 'MB.OFFICER')
    PASSWORD = os.getenv('APP_PASSWORD', '123456')
    OUTPUT_FILE = 'uiMap_selenium_fullrun_final_stats.xlsx'
    STATS_FILE = 'page_stats_final.xlsx'
    
    driver = setup_driver()
    
    # Load checkpoint and existing data
    checkpoint = load_checkpoint()
    processed_items_set = set(checkpoint.get('processed_items', []))
    start_index = checkpoint.get('last_index', 0)
    
    # Load existing data if resuming (check if Excel exists)
    all_data = []
    stats_data = []  # Track page statistics
    global_seen_rows = set()  # GLOBAL deduplication: track complete rows, not just XPaths
    
    if os.path.exists(OUTPUT_FILE):
        all_data = load_existing_data(OUTPUT_FILE)
        print(f"📂 Loaded {len(all_data)} existing records from {OUTPUT_FILE}")
        # Build global seen set from existing data - using ALL columns as key
        for item in all_data:
            row_key = (
                item.get('page', ''),
                item.get('relativeXpath', ''),
                item.get('elementName', ''),
                item.get('id', ''),
                item.get('name', ''),
                item.get('className', ''),
                item.get('tagName', ''),
                item.get('type', '')
            )
            global_seen_rows.add(row_key)
        print(f"   🔍 Tracking {len(global_seen_rows)} existing rows for deduplication")
    
    if start_index > 0:
        print(f"\n🔄 RESUMING from index {start_index} ({len(processed_items_set)} items already processed)\n")
    
    try:
        # Login
        login(driver, URL, USERNAME, PASSWORD)
        
        # Switch to menu frame
        if not get_menu_frame(driver, wait=True):
            print("❌ Failed to find menu frame")
            return
        
        # Find container
        container = driver.find_element(By.XPATH, "//div[starts-with(@id,'pane')]")
        print(f"📦 Container found: {container.get_attribute('id')}\n")
        
        # Auto-expand all menus
        expand_all_menus_recursive(driver)
        
        # Re-query after expansion and BUILD LINK LIST ONCE
        driver.switch_to.default_content()
        get_menu_frame(driver)  # No wait needed - already loaded
        container = driver.find_element(By.XPATH, "//div[starts-with(@id,'pane')]")
        
        # Find ALL leaf nodes with docommand
        all_leaves = container.find_elements(By.XPATH, ".//li[.//a[starts-with(@href,'javascript:docommand(')]]")
        
        # Build a simple list of link texts only (hierarchy extraction moved to click time)
        print(f"\n📋 Building menu item list...")
        items_to_process = []
        for leaf in all_leaves:
            if leaf.is_displayed():
                try:
                    link = leaf.find_element(By.XPATH, ".//a")
                    if link.is_displayed():
                        text = link.text.strip() or "Unknown"
                        # Store just text - get hierarchy when actually clicking
                        items_to_process.append(text)
                except:
                    pass
        
        total = len(items_to_process)
        print(f"🎯 Found {total} visible/clickable leaf nodes\n")
        
        # Process each item by finding it fresh each time (avoid stale element)
        for i in range(total):
            # Skip already processed items
            if i < start_index:
                continue
            
            text = items_to_process[i]
            
            print(f"[{i+1}/{total}] 🖱️ Clicking: {text}")
            
            # Re-query just this ONE link by text
            driver.switch_to.default_content()
            get_menu_frame(driver)  # No wait - frames already loaded
            
            try:
                # Find the specific link by its text
                link = driver.find_element(By.XPATH, f".//a[starts-with(@href='javascript:docommand(') and text()='{text}']")
            except:
                # Fallback: try finding by partial text if exact match fails
                try:
                    link = driver.find_element(By.XPATH, f".//a[starts-with(@href,'javascript:docommand(') and contains(text(),'{text[:20]}')]")
                except:
                    print(f"    ⚠️ Could not find link, skipping")
                    continue
            
            # Check if already processed (by page name only)
            if text in processed_items_set:
                print(f"    ⏭️ Already processed")
                continue
            
            # Scroll into view before clicking (minimal delay)
            try:
                driver.execute_script("arguments[0].scrollIntoView({behavior: 'auto', block: 'center'});", link)
            except:
                pass
            
            # Store original window handle
            main_window = driver.current_window_handle
            
            # Click the link with multiple fallback methods
            try:
                # Try regular click first
                try:
                    link.click()
                except:
                    # Try JavaScript click as fallback
                    driver.execute_script("arguments[0].click();", link)
                
                time.sleep(0.3)  # Reduced wait for popup
                
                # Check if popup opened
                windows = driver.window_handles
                if len(windows) > 1:
                    # Switch to popup
                    driver.switch_to.window(windows[-1])
                    # Removed verbose logging for speed
                    
                    time.sleep(0.2)  # Minimal wait for page load
                    
                    # Extract XPaths with GLOBAL row-level deduplication
                    extracted = extract_xpaths_from_page(driver, text, global_seen_rows)
                    all_data.extend(extracted)
                    stats_data.append({'page': text, 'count': len(extracted)})
                    print(f"    ✅ {len(extracted)} elements")  # Condensed output
                    
                    # Close popup
                    driver.close()
                    driver.switch_to.window(main_window)
                else:
                    print(f"    ⚠️ No popup opened")
                    stats_data.append({'page': text, 'count': 0})
                
                # Mark as processed and save checkpoint every 10 items
                processed_items_set.add(text)
                if (i + 1) % 10 == 0:
                    save_checkpoint(list(processed_items_set), i + 1)
                
                # Save Excel incrementally every 50 items
                if (i + 1) % 50 == 0:
                    try:
                        export_to_excel(all_data, OUTPUT_FILE)
                        export_stats_to_excel(stats_data, STATS_FILE)
                        print(f"    💾 Checkpoint: Saved {len(all_data)} elements to Excel")
                    except Exception as save_err:
                        print(f"    ⚠️ Incremental save failed: {save_err}")
                
            except Exception as e:
                print(f"    ❌ Error: {e}")
                # Even if error, mark as attempted and save checkpoint every 10 items
                processed_items_set.add(text)
                if (i + 1) % 10 == 0:
                    save_checkpoint(list(processed_items_set), i + 1)
                
                # Save Excel on error too (every 50 items)
                if (i + 1) % 50 == 0 and all_data:
                    try:
                        export_to_excel(all_data, OUTPUT_FILE)
                        export_stats_to_excel(stats_data, STATS_FILE)
                        print(f"    💾 Checkpoint: Saved {len(all_data)} elements to Excel (after error)")
                    except Exception as save_err:
                        print(f"    ⚠️ Incremental save failed: {save_err}")
                
                # Return to main window if stuck
                try:
                    driver.switch_to.window(main_window)
                except:
                    pass
        
        print(f"\n✨ Complete! Clicked {total} leaves, extracted {len(all_data)} elements")
        
    except KeyboardInterrupt:
        print(f"\n\n⚠️ Interrupted by user (Ctrl+C)")
        print(f"📊 Processed {len(all_data)} elements so far")
    except Exception as e:
        print(f"❌ Error during crawl: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        # Always export whatever data was collected
        if all_data:
            print(f"\n💾 Exporting {len(all_data)} elements to {OUTPUT_FILE}")
            try:
                export_to_excel(all_data, OUTPUT_FILE)
                export_stats_to_excel(stats_data, STATS_FILE)
                print(f"✅ Successfully exported to {OUTPUT_FILE}")
                print(f"✅ Successfully exported stats to {STATS_FILE}")
                
                # Clear checkpoint only if all items processed
                if len(processed_items_set) >= total:
                    clear_checkpoint()
                    print("✅ All items processed - checkpoint cleared")
            except Exception as export_error:
                print(f"❌ Export failed: {export_error}")
        else:
            print("\n⚠️ No data collected to export")
        
        print("\n🔚 Closing browser...")
        driver.quit()


if __name__ == '__main__':
    crawl_menu()

