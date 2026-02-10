from openpyxl import load_workbook

wb = load_workbook("menu_hierarchy2.xlsx")
ws = wb.active

print("Columns:", [c.value for c in ws[1]])
print(f"\nTotal rows: {ws.max_row}")
print("\nFirst 30 rows:")
for i in range(2, min(32, ws.max_row+1)):
    row = [str(c.value) if c.value else "NA" for c in ws[i]]
    print(f"Row {i}: NodeID={row[0]} | Parent={row[1][:25]} | Level={row[2]} | Text={row[3][:30]} | UniqueID={row[4][:40]}")

print("\n\nChecking level distribution:")
levels = {}
for i in range(2, ws.max_row+1):
    level = ws.cell(i, 3).value
    if level not in levels:
        levels[level] = 0
    levels[level] += 1
print(levels)

print("\n\nChecking for N/A unique IDs:")
na_count = 0
for i in range(2, min(100, ws.max_row+1)):
    unique_id = ws.cell(i, 5).value
    if unique_id == "N/A" or unique_id is None:
        node_id = ws.cell(i, 1).value
        text = ws.cell(i, 4).value
        level = ws.cell(i, 3).value
        parent = ws.cell(i, 2).value
        print(f"  Row {i}: NodeID={node_id}, Level={level}, Text={text}, Parent={parent}")
        na_count += 1
        if na_count >= 10:
            break

wb.close()
