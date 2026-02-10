"""
FAST Iframe-Aware Crawler - Optimized for speed
Target: <10 seconds per page (down from 3-4 minutes)

Speed optimizations:
- Removed all sleeps (use WebDriverWait only when needed)
- Removed duplicate element counting (analyze_page)
- Simplified XPath generation (no JavaScript)
- Batch saves every 100 items
- Direct element access without retries
"""

import time
import os
import json
from typing import List, Dict, Set
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl import Workbook, load_workbook

load_dotenv()

XPATH_OUTPUT_FILE = 'uiMap_fast.xlsx'
STATS_OUTPUT_FILE = 'page_stats_fast.xlsx'
ZERO_ELEMENTS_FILE = 'zero_elements_pages.xlsx'
CHECKPOINT_FILE = 'crawler_fast_checkpoint.json'
MAX_RETRIES = 2  # Retry failed pages up to 2 times

from crawler import setup_driver, login, get_menu_frame


def expand_all_menus_fast(driver):
    """Expand all menus - FAST version without sleeps."""
    print("🔧 Auto-expanding all menu items...")
    driver.switch_to.default_content()
    get_menu_frame(driver)
    
    expanded_total = 0
    max_passes = 20
    
    for pass_num in range(max_passes):
        try:
            container = driver.find_element(By.XPATH, "//div[starts-with(@id,'pane')]")
            expand_icons = container.find_elements(By.XPATH, ".//span[contains(@onclick, 'ProcessMouseClick')]")
            
            if not expand_icons:
                break
            
            expanded_this_pass = 0
            for icon in expand_icons:
                try:
                    parent_li = icon.find_element(By.XPATH, "./ancestor::li[1]")
                    child_uls = parent_li.find_elements(By.XPATH, "./ul")
                    
                    if child_uls and not any(ul.is_displayed() for ul in child_uls):
                        icon.click()
                        expanded_this_pass += 1
                        # NO SLEEP - let browser handle it
                except:
                    continue
            
            expanded_total += expanded_this_pass
            print(f"   Pass {pass_num + 1}: Expanded {expanded_this_pass} nodes (Total: {expanded_total})")
            
            if expanded_this_pass == 0:
                print(f"✅ All menus fully expanded")
                break
        except:
            break
    
    return expanded_total


def extract_xpaths_fast(driver, page_name: str, seen_rows: Set) -> tuple:
    """
    Fast extraction from main document + iframes.
    Returns: (extracted_list, stats_dict)
    """
    extracted = []
    stats = {
        'page': page_name,
        'main_count': 0,
        'iframe_count': 0,
        'total_iframes': 0,
        'extraction_time_ms': 0
    }
    
    start_time = time.time()
    
    # Simpler selector - just the essentials
    selector = 'input, select, textarea'
    
    def extract_from_context(context_name="main"):
        """Extract from current context."""
        found = []
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, selector)
            
            for idx, elem in enumerate(elements):
                try:
                    tag = elem.tag_name
                    elem_id = elem.get_attribute('id') or ''
                    elem_name = elem.get_attribute('name') or ''
                    elem_type = elem.get_attribute('type') or ''
                    
                    # Skip hidden inputs - they're not fillable
                    if tag == 'input' and elem_type == 'hidden':
                        continue
                    
                    # Skip if no identifier
                    if not elem_id and not elem_name:
                        continue
                    
                    # Simple XPath generation
                    if elem_id:
                        xpath = f"//{tag}[@id='{elem_id}']"
                    elif elem_name:
                        xpath = f"//{tag}[@name='{elem_name}']"
                    else:
                        continue
                    
                    # Unique key
                    key = (page_name, xpath, elem_id, elem_name, tag, elem_type)
                    
                    if key not in seen_rows:
                        seen_rows.add(key)
                        found.append({
                            'page': page_name,
                            'xpath': xpath,
                            'id': elem_id,
                            'name': elem_name,
                            'tag': tag,
                            'type': elem_type,
                            'context': context_name
                        })
                except:
                    continue
        except:
            pass
        
        return found
    
    # Extract from main
    main_results = extract_from_context("main")
    extracted.extend(main_results)
    stats['main_count'] = len(main_results)
    
    # Extract from iframes
    try:
        iframes = driver.find_elements(By.TAG_NAME, 'iframe')
        stats['total_iframes'] = len(iframes)
        
        for idx, iframe in enumerate(iframes):
            try:
                driver.switch_to.frame(iframe)
                iframe_results = extract_from_context(f"iframe_{idx}")
                extracted.extend(iframe_results)
                stats['iframe_count'] += len(iframe_results)
                driver.switch_to.default_content()
            except:
                try:
                    driver.switch_to.default_content()
                except:
                    pass
    except:
        pass
    
    stats['extraction_time_ms'] = int((time.time() - start_time) * 1000)
    
    return extracted, stats


def load_checkpoint():
    """Load checkpoint if exists."""
    if os.path.exists(CHECKPOINT_FILE):
        try:
            with open(CHECKPOINT_FILE, 'r') as f:
                return json.load(f)
        except:
            return None
    return None


def save_checkpoint(page_index, page_name, total):
    """Save current progress."""
    checkpoint = {
        'page_index': page_index,
        'page_name': page_name,
        'total': total,
        'timestamp': time.strftime('%Y-%m-%d %H:%M:%S')
    }
    with open(CHECKPOINT_FILE, 'w') as f:
        json.dump(checkpoint, f, indent=2)


def clear_checkpoint():
    """Remove checkpoint file."""
    if os.path.exists(CHECKPOINT_FILE):
        os.remove(CHECKPOINT_FILE)


def initialize_files(resume=False):
    """Create or load Excel files."""
    if resume and os.path.exists(XPATH_OUTPUT_FILE):
        # Load existing workbooks
        xpath_wb = load_workbook(XPATH_OUTPUT_FILE)
        xpath_ws = xpath_wb.active
        stats_wb = load_workbook(STATS_OUTPUT_FILE)
        stats_ws = stats_wb.active
        zero_wb = load_workbook(ZERO_ELEMENTS_FILE) if os.path.exists(ZERO_ELEMENTS_FILE) else Workbook()
        zero_ws = zero_wb.active
    else:
        # Create new workbooks
        xpath_wb = Workbook()
        xpath_ws = xpath_wb.active
        xpath_ws.title = 'XPaths'
        xpath_ws.append(['page', 'xpath', 'id', 'name', 'tag', 'type', 'context'])
        
        stats_wb = Workbook()
        stats_ws = stats_wb.active
        stats_ws.title = 'Stats'
        stats_ws.append(['page', 'elements', 'main', 'iframes_found', 'iframes_extracted', 'time_ms', 'timestamp'])
        
        zero_wb = Workbook()
        zero_ws = zero_wb.active
        zero_ws.title = 'Zero Elements'
        zero_ws.append(['page', 'iframes_found', 'time_ms', 'timestamp', 'reason'])
    
    return xpath_wb, xpath_ws, stats_wb, stats_ws, zero_wb, zero_ws


def crawl_fast():
    """Fast crawler - optimized for speed with checkpoint/resume support."""
    print("="*70)
    
    # Check for checkpoint
    checkpoint = load_checkpoint()
    resume = checkpoint is not None
    start_index = checkpoint['page_index'] if resume else 0
    
    if resume:
        print(f"📍 Resuming from checkpoint: page {start_index + 1} ({checkpoint['page_name']})")
        print(f"   Last run: {checkpoint['timestamp']}\n")
    
    xpath_wb, xpath_ws, stats_wb, stats_ws, zero_wb, zero_ws = initialize_files(resume)
    
    seen_rows = set()
    xpath_count = 0
    zero_count = 0
    
    driver = setup_driver()
    
    try:
        # Login
        print("\n🔐 Logging in...")
        url = os.getenv('T24_URL')
        username = os.getenv('T24_USERNAME')
        password = os.getenv('T24_PASSWORD')
        login(driver, url, username, password)
        
        # Expand menu
        print("🔧 Expanding menu...")
        expand_all_menus_fast(driver)
        
        # Wait for menu to be ready
        get_menu_frame(driver, wait=True)
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, ".//li[.//a[starts-with(@href,'javascript:docommand(')]]"))
        )
        
        # Get all menu items
        all_leaves = driver.find_elements(By.XPATH, ".//li[.//a[starts-with(@href,'javascript:docommand(')]]")
        
        items = []
        for leaf in all_leaves:
            if leaf.is_displayed():
                try:
                    link = leaf.find_element(By.XPATH, ".//a")
                    if link.is_displayed():
                        items.append(link.text.strip() or "Unknown")
                except:
                    pass
        
        # Check for duplicates before deduplication
        original_count = len(items)
        duplicates = {}
        for item in items:
            duplicates[item] = duplicates.get(item, 0) + 1
        
        duplicate_pages = {k: v for k, v in duplicates.items() if v > 1}
        if duplicate_pages:
            print(f"⚠️ Found duplicate menu entries:")
            for page, count in sorted(duplicate_pages.items(), key=lambda x: -x[1])[:10]:
                print(f"   '{page}' appears {count} times")
        
        # Deduplicate items list - process each unique page only once
        items = list(dict.fromkeys(items))  # Preserves order, removes duplicates
        
        total = len(items)
        print(f"🎯 Processing {total} unique pages (from {original_count} total menu items)")
        
        if resume:
            print(f"⏩ Skipping first {start_index} pages (already processed)\n")
        else:
            print()
        
        start_time = time.time()
        
        # Process pages (skip to checkpoint if resuming)
        for i in range(start_index, total):
            page_start = time.time()
            page_name = items[i]
            
            print(f"[{i+1}/{total}] {page_name}")
            
            # Try to process page, restart browser if crashed
            success = False
            attempt = 0
            
            while attempt <= MAX_RETRIES and not success:
                try:
                    # Check if browser is alive
                    try:
                        _ = driver.current_window_handle
                    except:
                        # Browser crashed, restart it
                        print(f"  🔄 Browser crashed, restarting...")
                        try:
                            driver.quit()
                        except:
                            pass
                        
                        driver = setup_driver()
                        login(driver, url, username, password)
                        expand_all_menus_fast(driver)
                        print(f"  ✅ Browser restarted, continuing...")
                    
                    # Re-find link
                    driver.switch_to.default_content()
                    get_menu_frame(driver)
                    
                    try:
                        link = driver.find_element(By.XPATH, f".//a[starts-with(@href,'javascript:docommand(') and text()='{page_name}']")
                    except:
                        try:
                            link = driver.find_element(By.XPATH, f".//a[starts-with(@href,'javascript:docommand(') and contains(text(),'{page_name[:20]}')]")
                        except:
                            print(f"  ⚠️ Link not found, skipping...")
                            success = True  # Mark as success to move to next page
                            break
                    
                    main_window = driver.current_window_handle
                    
                    # Click and wait for popup
                    link.click()
                    
                    # Wait for popup window (max 2 seconds)
                    WebDriverWait(driver, 2).until(lambda d: len(d.window_handles) > 1)
                    
                    # Switch to popup
                    driver.switch_to.window(driver.window_handles[-1])
                    
                    # Wait for page body
                    WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
                    
                    # Quick check if page has any inputs at all
                    try:
                        WebDriverWait(driver, 1).until(
                            lambda d: len(d.find_elements(By.CSS_SELECTOR, 'input, select, textarea')) > 0
                        )
                        # If inputs exist, wait a bit more for dealbox fields to load
                        try:
                            WebDriverWait(driver, 2).until(
                                lambda d: len(d.find_elements(By.CSS_SELECTOR, '.dealbox, [id^="fieldName:"]')) > 0
                            )
                        except:
                            pass  # Inputs exist but not dealbox type, proceed anyway
                    except:
                        pass  # No inputs at all, skip waiting
                    
                    # Extract XPaths
                    extracted, stats = extract_xpaths_fast(driver, page_name, seen_rows)
                    
                    # Save XPath rows
                    for xp in extracted:
                        xpath_ws.append([xp['page'], xp['xpath'], xp['id'], xp['name'], xp['tag'], xp['type'], xp['context']])
                        xpath_count += 1
                    
                    # Save stats row
                    stats_ws.append([
                        page_name,
                        len(extracted),
                        stats['main_count'],
                        stats['total_iframes'],
                        stats['iframe_count'],
                        stats['extraction_time_ms'],
                        time.strftime('%H:%M:%S')
                    ])
                    
                    # Track zero-element pages separately
                    if len(extracted) == 0:
                        zero_count += 1
                        reason = "Duplicate menu entry (already extracted)" if page_name in [r[0] for r in zero_ws.iter_rows(min_row=2, values_only=True)] else "No fillable fields found"
                        zero_ws.append([
                            page_name,
                            stats['total_iframes'],
                            stats['extraction_time_ms'],
                            time.strftime('%H:%M:%S'),
                            reason
                        ])
                    
                    # Close popup
                    driver.close()
                    driver.switch_to.window(main_window)
                    
                    success = True
                    
                except Exception as e:
                    attempt += 1
                    error_msg = str(e)[:60]
                    
                    if attempt <= MAX_RETRIES:
                        print(f"  ⚠️ Attempt {attempt} failed: {error_msg}")
                    else:
                        print(f"  ❌ Failed after {MAX_RETRIES} attempts, skipping page")
                    
                    # Clean up any open windows
                    try:
                        if len(driver.window_handles) > 1:
                            driver.switch_to.window(driver.window_handles[-1])
                            driver.close()
                        driver.switch_to.window(main_window)
                    except:
                        pass
            
            # Save checkpoint after each page
            save_checkpoint(i, page_name, total)
            
            # Save files every 100 items
            if (i + 1) % 100 == 0:
                xpath_wb.save(XPATH_OUTPUT_FILE)
                stats_wb.save(STATS_OUTPUT_FILE)
                zero_wb.save(ZERO_ELEMENTS_FILE)
                elapsed = int(time.time() - start_time)
                pages_done = i + 1 - start_index
                avg_per_page = elapsed / pages_done if pages_done > 0 else 0
                remaining = int((total - i - 1) * avg_per_page / 60) if avg_per_page > 0 else 0
                print(f"  💾 Progress: {i+1}/{total} | Avg: {avg_per_page:.1f}s/page | ETA: {remaining}min")
        
        # Final save
        xpath_wb.save(XPATH_OUTPUT_FILE)
        stats_wb.save(STATS_OUTPUT_FILE)
        zero_wb.save(ZERO_ELEMENTS_FILE)
        clear_checkpoint()
        
        total_time = int(time.time() - start_time)
        pages_done = total - start_index
        avg_time = total_time / pages_done if pages_done > 0 else 0
        
        print(f"\n✨ Complete!")
        print(f"   Total: {xpath_count} XPaths from {total} pages")
        print(f"   Zero-element pages: {zero_count}")
        print(f"   Time: {total_time}s ({avg_time:.1f}s per page)")
        print(f"\n📊 Files:")
        print(f"   1. {XPATH_OUTPUT_FILE} - XPath records")
        print(f"   2. {STATS_OUTPUT_FILE} - All page statistics")
        print(f"   3. {ZERO_ELEMENTS_FILE} - Pages with 0 elements")
        
    except KeyboardInterrupt:
        print("\n\n⏸️ Interrupted - Saving...")
        xpath_wb.save(XPATH_OUTPUT_FILE)
        stats_wb.save(STATS_OUTPUT_FILE)
        zero_wb.save(ZERO_ELEMENTS_FILE)
        print(f"💾 Checkpoint saved. Run again to resume from page {xpath_count}")
        
    except Exception as e:
        print(f"\n❌ Fatal: {e}")
        import traceback
        traceback.print_exc()
        
    finally:
        try:
            driver.quit()
        except:
            pass


if __name__ == "__main__":
    crawl_fast()
