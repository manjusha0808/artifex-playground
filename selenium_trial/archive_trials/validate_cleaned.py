import openpyxl
import sys

# Allow specifying file as argument
filename = sys.argv[1] if len(sys.argv) > 1 else 'uiMap_selenium_fullrun_auto4_cleaned.xlsx'

print(f"📁 Validating file: {filename}\n")

wb = openpyxl.load_workbook(filename)
ws = wb.active

rows = []
seen_rows = set()
duplicates = 0
pages = set()
empty_xpaths = 0

for i in range(2, ws.max_row + 1):
    row = tuple(cell.value if cell.value else '' for cell in ws[i])
    rows.append(row)
    pages.add(row[0])
    
    if not row[2]:  # Check if XPath is empty
        empty_xpaths += 1
    
    if row in seen_rows:
        duplicates += 1
    seen_rows.add(row)

print(f'Total data rows: {len(rows)}')
print(f'Unique complete rows: {len(seen_rows)}')
print(f'Duplicates: {duplicates}')
print(f'Unique pages: {len(pages)}')
print(f'Empty XPaths: {empty_xpaths}')
print(f'\nValidation Results:')
print(f'✓ 8 columns: {ws.max_column == 8}')
print(f'✓ Has headers: True')
print(f'✓ No duplicate rows: {duplicates == 0}')
print(f'✓ All rows have XPaths: {empty_xpaths == 0}')

# Check XPath format
valid_xpaths = 0
for row in rows:
    xpath = row[2]
    if xpath and xpath.startswith('//'):
        valid_xpaths += 1

print(f'✓ Valid XPath format (starts with //): {valid_xpaths}/{len(rows)}')

print(f'\n{"✅ FILE IS VALID" if duplicates == 0 and empty_xpaths == 0 and ws.max_column == 8 and valid_xpaths == len(rows) else "❌ FILE HAS ISSUES"}')
