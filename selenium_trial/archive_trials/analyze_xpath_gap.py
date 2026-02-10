"""
Analyze actual XPath vs extracted XPath for N/A nodes
Find the disconnect and generate position-based selectors
"""

from openpyxl import load_workbook

def analyze_xpath_mismatch():
    """
    Compare what we extracted vs what the browser actually has
    """
    
    # Load the file
    excel_file = 'menu_hierarchy3.xlsx'
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Find the specific node
    search_text = 'Joint Customer Relationship Enquiry'
    
    print('=== ANALYZING XPATH MISMATCH ===\n')
    print(f'Searching for: {search_text}\n')
    
    # Get headers
    headers = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        headers[header] = col
    
    # Find all matching rows
    found = False
    for row in range(2, ws.max_row + 1):
        node_text = ws.cell(row, headers['Node Text']).value
        if node_text == search_text:
            found = True
            print(f'✓ Found node at row {row}')
            print(f'  Node Text: {node_text}')
            print(f'  Full Path: {ws.cell(row, headers["Full Path"]).value}')
            print(f'  Level: {ws.cell(row, headers["Level"]).value}')
            print(f'  Type: {ws.cell(row, headers["Type"]).value}')
            print(f'  Unique ID: {ws.cell(row, headers["Unique ID"]).value}')
            print(f'  Parent Node: {ws.cell(row, headers["Parent Node"]).value}')
            print(f'  XPath (Unique) in Excel: {ws.cell(row, headers["XPath (Unique)"]).value}')
            print()
    
    if not found:
        print(f'❌ Node "{search_text}" not found in Excel\n')
    
    print('=== THE PROBLEM ===\n')
    print(f'Your actual XPath from browser: //*[@id="pane_"]/ul[1]/li/ul/li[1]/ul/li[6]/ul/li[6]/a')
    print(f'  → This is ABSOLUTE position-based')
    print(f'  → Starts from pane_ element')
    print(f'  → Uses index positions like li[6]')
    print()
    print(f'My generated XPath: //li//span[contains(text()...)]')
    print(f'  → This is RELATIVE text-based')
    print(f'  → Assumes span with onclick')
    print(f'  → MISMATCH: The actual element is <a>, not <span>')
    print()
    
    print('=== WHAT WENT WRONG ===\n')
    print('The extractor looks for:')
    print('  1. <span> with ProcessMouseClick onclick → marked as parent')
    print('  2. <a> with docommand href → marked as leaf')
    print()
    print('But: "Joint Customer Relationship Enquiry" might be a LEAF with clickable <a>')
    print('     Yet the extractor found a <span> first and marked it as parent')
    print('     So it classified as Parent (Expandable) with N/A XPath')
    print()
    
    print('=== SOLUTION ===\n')
    print('Need to re-examine the extraction logic:')
    print('  - Maybe <a> tag comes AFTER <span> in DOM order')
    print('  - Maybe the text appears in both, and span is found first')
    print('  - Need position-based XPath from absolute paths in browser')
    print()
    
    print('=== BETTER APPROACH ===\n')
    print('Generate XPath from your browser path:')
    print()
    print('Your path: //*[@id="pane_"]/ul[1]/li/ul/li[1]/ul/li[6]/ul/li[6]/a')
    print()
    print('This means:')
    print('  - Element ID: pane_')
    print('  - 6 levels deep with li/ul alternating')
    print('  - Final position: 6th li child')
    print('  - Final element: <a> tag')
    print()
    print('Can be simplified to:')
    print('  //*[@id="pane_"]//li[6]//a (if unique position)')
    print('  or')
    print("  //*[@id=\"pane_\"]//span[text()='Joint Customer Relationship Enquiry']/..//a")
    print()

if __name__ == '__main__':
    analyze_xpath_mismatch()
