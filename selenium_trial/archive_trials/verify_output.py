"""Verification script to validate the crawler output Excel file."""

import openpyxl
from collections import defaultdict
import sys

def verify_excel_output(filepath: str = 'uiMap_selenium_fullrun_modified.xlsx'):
    """Verify the crawler output for duplicates and integrity."""
    
    print(f"📋 Verifying: {filepath}\n")
    
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        print(f"✅ Headers found: {headers}\n")
        
        # Expected headers
        expected_headers = ['hierarchy', 'pageName', 'elementName', 'relativeXpath', 'elementId', 'elementNameAttr', 'className', 'tagName', 'inputType']
        
        if headers != expected_headers:
            print(f"⚠️  Warning: Headers don't match expected format")
            print(f"   Expected: {expected_headers}")
            print(f"   Found:    {headers}\n")
        
        # Collect data
        total_rows = 0
        xpath_count = defaultdict(int)  # Count occurrences of each xpath
        page_xpath_map = defaultdict(list)  # Map page to xpaths
        xpath_page_map = defaultdict(set)  # Map xpath to pages
        page_field_count = defaultdict(int)  # Count fields per page
        duplicate_xpaths = []
        
        # Process rows (skip header)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            total_rows += 1
            hierarchy, page_name, element_name, xpath, elem_id, elem_name_attr, class_name, tag_name, input_type = row
            
            # Track xpath occurrences
            if xpath:
                xpath_count[xpath] += 1
                page_xpath_map[page_name].append(xpath)
                xpath_page_map[xpath].add(page_name)
                page_field_count[page_name] += 1
                
                if xpath_count[xpath] > 1:
                    if xpath not in [dup[0] for dup in duplicate_xpaths]:
                        duplicate_xpaths.append((xpath, page_name, row_num))
        
        # Report findings
        print(f"📊 Total Records: {total_rows}")
        print(f"📊 Unique XPaths: {len(xpath_count)}")
        print(f"📊 Unique Pages: {len(page_field_count)}\n")
        
        # Check for duplicates
        duplicate_count = sum(1 for count in xpath_count.values() if count > 1)
        
        if duplicate_count > 0:
            print(f"⚠️  DUPLICATES FOUND: {duplicate_count} XPaths appear multiple times\n")
            
            # Show detailed duplicate analysis
            print("🔍 Duplicate Analysis:")
            for xpath, count in sorted(xpath_count.items(), key=lambda x: x[1], reverse=True):
                if count > 1:
                    pages = xpath_page_map[xpath]
                    if len(pages) == 1:
                        # Same xpath appears multiple times for SAME page
                        print(f"   ❌ DUPLICATE: '{xpath}'")
                        print(f"      Appears {count} times on SAME page: {list(pages)[0]}")
                    else:
                        # Same xpath appears on DIFFERENT pages (might be valid)
                        print(f"   ⚠️  SHARED: '{xpath}'")
                        print(f"      Appears on {len(pages)} different pages: {list(pages)[:3]}...")
            print()
        else:
            print("✅ NO DUPLICATES: All XPaths are unique\n")
        
        # Check page-xpath relationship (Column B and Column D)
        print("🔍 Page-XPath Relationship Verification:")
        relationship_issues = 0
        
        for page_name, xpaths in page_xpath_map.items():
            unique_xpaths = set(xpaths)
            if len(xpaths) != len(unique_xpaths):
                # Same page has duplicate xpaths
                duplicates_on_page = len(xpaths) - len(unique_xpaths)
                print(f"   ❌ Page '{page_name}' has {duplicates_on_page} duplicate XPaths")
                relationship_issues += 1
        
        if relationship_issues == 0:
            print("   ✅ All pages have unique XPaths (no duplicates within same page)\n")
        else:
            print(f"   ⚠️  Found {relationship_issues} pages with duplicate XPaths\n")
        
        # Field count per page
        print("📈 Fields Per Page Statistics:")
        field_counts = list(page_field_count.values())
        if field_counts:
            print(f"   Min fields: {min(field_counts)}")
            print(f"   Max fields: {max(field_counts)}")
            print(f"   Avg fields: {sum(field_counts) / len(field_counts):.1f}\n")
        
        # Show pages with most fields
        print("📋 Top 10 Pages by Field Count:")
        top_pages = sorted(page_field_count.items(), key=lambda x: x[1], reverse=True)[:10]
        for idx, (page, count) in enumerate(top_pages, 1):
            print(f"   {idx}. {page}: {count} fields")
        print()
        
        # Show pages with duplicate fields
        pages_with_dupes = []
        for page_name, xpaths in page_xpath_map.items():
            unique = set(xpaths)
            if len(xpaths) != len(unique):
                pages_with_dupes.append((page_name, len(xpaths), len(unique)))
        
        if pages_with_dupes:
            print("⚠️  Pages with Duplicate Fields:")
            for page, total, unique in sorted(pages_with_dupes, key=lambda x: x[1]-x[2], reverse=True)[:10]:
                print(f"   - {page}: {total} total, {unique} unique, {total-unique} duplicates")
            print()
        
        # Final verdict
        print("="*60)
        if duplicate_count == 0 and relationship_issues == 0:
            print("✅ VALIDATION PASSED: Output is valid and reliable")
        elif relationship_issues > 0:
            print("❌ VALIDATION FAILED: Double crawling detected!")
            print("   Same fields are being extracted multiple times from same pages")
        else:
            print("⚠️  VALIDATION WARNING: Some issues detected")
        print("="*60)
        
        return duplicate_count, relationship_issues
        
    except FileNotFoundError:
        print(f"❌ Error: File '{filepath}' not found")
        print(f"   Make sure the crawler has run and created the output file")
        return None, None
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return None, None


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description="Verify crawler output Excel file")
    parser.add_argument('--file', default='uiMap_selenium_fullrun_modified.xlsx', help='Path to Excel file')
    
    args = parser.parse_args()
    
    duplicates, issues = verify_excel_output(args.file)
    
    # Exit with error code if issues found
    if duplicates is not None and issues is not None:
        if issues > 0:
            sys.exit(1)  # Exit with error if relationship issues found
        elif duplicates > 0:
            sys.exit(2)  # Exit with warning if only duplicates found
