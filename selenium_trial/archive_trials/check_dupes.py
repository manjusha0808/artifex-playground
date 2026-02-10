"""Check duplicate details"""
from openpyxl import load_workbook

wb = load_workbook('uiMap_selenium_fullrun_auto3.xlsx')
ws = wb.active

print("Analyzing duplicates...")
xpath_details = {}

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    if not row[3]:
        continue
    xpath = row[3]
    if xpath not in xpath_details:
        xpath_details[xpath] = []
    xpath_details[xpath].append({
        'row': i,
        'hierarchy': row[0],
        'page': row[1],
        'element': row[2]
    })

duplicates = {k: v for k, v in xpath_details.items() if len(v) > 1}

print(f"\nTotal duplicate XPaths: {len(duplicates)}")
print(f"Total unique XPaths: {len(xpath_details) - len(duplicates)}")

print("\n\nFirst 3 duplicate examples:")
for i, (xpath, occurrences) in enumerate(list(duplicates.items())[:3], 1):
    print(f"\n{i}. XPath: {xpath}")
    print(f"   Appears {len(occurrences)} times:")
    for occ in occurrences:
        print(f"      Row {occ['row']}: {occ['element']}")
