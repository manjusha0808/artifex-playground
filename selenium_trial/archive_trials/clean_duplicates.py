"""
Remove duplicate XPaths from the Excel file, keeping only the first occurrence.
"""
from openpyxl import load_workbook

filepath = 'uiMap_selenium_fullrun_auto3.xlsx'

print(f"📊 Cleaning duplicates from: {filepath}")

wb = load_workbook(filepath)
ws = wb.active

# Track seen XPaths and rows to delete
seen_xpaths = set()
rows_to_delete = []

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
    xpath = row[3].value  # relativeXpath column
    if xpath:
        if xpath in seen_xpaths:
            rows_to_delete.append(i)
        else:
            seen_xpaths.add(xpath)

print(f"   Found {len(rows_to_delete)} duplicate rows to remove")
print(f"   Keeping {len(seen_xpaths)} unique XPaths")

# Delete rows in reverse order (so indices don't shift)
for row_idx in reversed(rows_to_delete):
    ws.delete_rows(row_idx, 1)

# Save cleaned file
output_file = 'uiMap_selenium_fullrun_auto3_cleaned.xlsx'
wb.save(output_file)
print(f"✅ Saved cleaned file: {output_file}")
print(f"   Total rows now: {ws.max_row - 1}")
