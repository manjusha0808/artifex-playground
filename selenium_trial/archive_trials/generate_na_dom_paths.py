"""
Generate DOM Paths for missing XPath nodes (Parent nodes with N/A)
Creates Inspector-searchable selectors for each parent node without XPath
Output: Excel file with alternative selectors for DevTools inspection
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def generate_dom_paths_for_na_nodes():
    """
    Read menu_hierarchy3.xlsx and generate DOM paths only for nodes with N/A XPath
    """
    
    # Load the Excel file
    excel_file = 'menu_hierarchy3.xlsx'
    wb = load_workbook(excel_file)
    ws = wb.active
    
    print('=== GENERATING DOM PATHS FOR N/A XPATH NODES ===\n')
    
    # Get header row
    headers = {}
    for col in range(1, ws.max_column + 1):
        header_name = ws.cell(1, col).value
        headers[header_name] = col
    
    # Extract columns we need
    xpath_col = headers.get('XPath (Unique)')
    full_path_col = headers.get('Full Path')
    node_text_col = headers.get('Node Text')
    parent_col = headers.get('Parent Node')
    level_col = headers.get('Level')
    node_id_col = headers.get('Node ID')
    unique_id_col = headers.get('Unique ID')
    
    # Collect all rows and find N/A rows
    all_rows = []
    na_rows = []
    
    for row in range(2, ws.max_row + 1):
        row_data = {
            'Node ID': ws.cell(row, node_id_col).value,
            'Full Path': ws.cell(row, full_path_col).value,
            'Node Text': ws.cell(row, node_text_col).value,
            'Parent Node': ws.cell(row, parent_col).value,
            'Level': ws.cell(row, level_col).value,
            'XPath (Unique)': ws.cell(row, xpath_col).value,
            'Unique ID': ws.cell(row, unique_id_col).value,
        }
        all_rows.append(row_data)
        
        # Check if this row has N/A xpath
        if row_data['XPath (Unique)'] == 'N/A':
            na_rows.append(row_data)
    
    print(f'Found {len(na_rows)} parent nodes with N/A XPath\n')
    
    # Generate alternative selectors for each
    selectors = []
    
    for row_data in na_rows:
        node_id = row_data['Node ID']
        full_path = row_data['Full Path']
        node_text = row_data['Node Text']
        parent_node = row_data['Parent Node']
        level = row_data['Level']
        unique_id = row_data['Unique ID']
        node_id = row_data['Node ID']
        full_path = row_data['Full Path']
        node_text = row_data['Node Text']
        parent_node = row_data['Parent Node']
        level = row_data['Level']
        unique_id = row_data['Unique ID']
        
        # Method 1: Text-based XPath (simplest for DevTools)
        xpath_text = f"//span[contains(text(), '{node_text}')]"
        
        # Method 2: More specific - span with parent context
        # Looking for span inside li that contains this text and has ProcessMouseClick onclick
        xpath_specific = f"//li//span[contains(text(), '{node_text}') and contains(@onclick, 'ProcessMouseClick')]"
        
        # Method 3: Hierarchical - using full path to narrow down
        # Build path-based selector
        path_parts = str(full_path).split(' > ')
        dom_breadcrumb = ' > '.join([f'li[text*="{part.strip()}"]' for part in path_parts])
        
        # Method 4: CSS-like selector (text matching with parent)
        css_selector = f"span:has-text('{node_text}')[onclick*='ProcessMouseClick']"
        
        # Method 5: Position + parent based (more reliable)
        # Count siblings with same parent to get position
        siblings_with_na = [r for r in na_rows if r['Parent Node'] == parent_node and r['Level'] == level]
        position_in_parent = len([s for s in siblings_with_na if str(s['Node Text']) <= str(node_text)])
        
        # DOM path that can be used in DevTools
        dom_path = f"Find in: Parent '{parent_node}' > Child position {position_in_parent} > span (text: '{node_text}')"
        
        # Create inspector-ready selector
        inspector_path = f"{parent_node} → {node_text}"
        
        selectors.append({
            'Node ID': node_id,
            'Full Path': full_path,
            'Node Text': node_text,
            'Parent Node': parent_node,
            'Level': level,
            'Unique ID': unique_id,
            'XPath (Text)': xpath_text,
            'XPath (Specific)': xpath_specific,
            'DOM Breadcrumb': dom_breadcrumb,
            'Inspector Path': inspector_path,
            'Position in Parent': position_in_parent,
        })
    
    print(f'Generated selectors for {len(selectors)} nodes\n')
    
    # Export results to Excel
    output_file = 'na_nodes_dom_paths.xlsx'
    from openpyxl import Workbook
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'NA Node Selectors'
    
    # Write headers
    headers_list = ['Node ID', 'Full Path', 'Node Text', 'Parent Node', 'Level', 'Unique ID', 
                   'XPath (Text)', 'XPath (Specific)', 'DOM Breadcrumb', 'Inspector Path', 'Position in Parent']
    out_ws.append(headers_list)
    
    # Write data
    for selector in selectors:
        row = [
            selector['Node ID'],
            selector['Full Path'],
            selector['Node Text'],
            selector['Parent Node'],
            selector['Level'],
            selector['Unique ID'],
            selector['XPath (Text)'],
            selector['XPath (Specific)'],
            selector['DOM Breadcrumb'],
            selector['Inspector Path'],
            selector['Position in Parent'],
        ]
        out_ws.append(row)
    
    # Auto-adjust column widths
    for col in range(1, len(headers_list) + 1):
        max_length = len(str(headers_list[col-1]))
        col_letter = get_column_letter(col)
        for row in range(2, len(selectors) + 2):
            cell_value = out_ws.cell(row, col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 150)
        out_ws.column_dimensions[col_letter].width = adjusted_width
    
    out_wb.save(output_file)
    print(f'✅ Saved to: {output_file}\n')
    
    # Print samples
    print('=== SAMPLE SELECTORS (First 5 nodes) ===\n')
    for i, selector in enumerate(selectors[:5]):
        print(f"Node {i+1}: {selector['Node Text']}")
        print(f"  Parent: {selector['Parent Node']}")
        print(f"  Full Path: {selector['Full Path']}")
        print(f"  → For DevTools Inspector search:")
        print(f"     Copy this XPath: {selector['XPath (Specific)']}")
        print(f"     Or: {selector['Inspector Path']}")
        print()
    
    # Create a summary for each N/A node showing HOW to find it in DevTools
    print('\n=== HOW TO USE IN DEVTOOLS ===\n')
    print('1. Open DevTools (F12)')
    print('2. Go to Inspector tab')
    print('3. Use Ctrl+F to open search')
    print('4. Copy and paste the XPath from "XPath (Specific)" column')
    print('5. Or manually navigate: Find parent span with "ProcessMouseClick" onclick\n')
    
    print('=== QUICK REFERENCE: All N/A Nodes ===\n')
    for i, selector in enumerate(selectors):
        print(f"{i+1}. {selector['Node Text']:<50} → Parent: {selector['Parent Node']:<40}")
    
    return selectors

if __name__ == '__main__':
    result_df = generate_dom_paths_for_na_nodes()
    print(f'\n✅ Complete! Check na_nodes_dom_paths.xlsx for full selectors')
