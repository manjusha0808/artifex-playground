from openpyxl import load_workbook
from collections import defaultdict

wb = load_workbook('menu_hierarchy3.xlsx')
ws = wb.active

max_row = ws.max_row
max_col = ws.max_column

# Get header
header = []
for col in range(1, max_col + 1):
    header.append(str(ws.cell(1, col).value))

print('=== ANALYZING "N/A" STRING PATTERNS IN EXCEL ===')
print()

# Count "N/A" by column
print('Columns with "N/A" values:')
na_string_counts = {}
for col in range(1, max_col + 1):
    col_name = header[col-1]
    na_count = 0
    for row in range(2, max_row + 1):
        val = ws.cell(row, col).value
        if val == "N/A" or str(val).strip() == "N/A":
            na_count += 1
    if na_count > 0:
        pct = (na_count / (max_row - 1) * 100)
        na_string_counts[col_name] = na_count
        print(f'  {col_name}: {na_count} rows ({pct:.2f}%)')

print()
print('=== DETAILED N/A ANALYSIS ===')

# Find all rows with N/A in XPath column
xpath_col = None
for col in range(1, max_col + 1):
    if header[col-1] == "XPath (Unique)":
        xpath_col = col
        break

if xpath_col:
    print(f'Rows with "N/A" in XPath (Unique) column:')
    xpath_na_rows = []
    for row in range(2, max_row + 1):
        val = ws.cell(row, xpath_col).value
        if val == "N/A":
            xpath_na_rows.append(row)
    
    print(f'Total: {len(xpath_na_rows)} rows')
    print()
    print('First 20 examples:')
    for row in xpath_na_rows[:20]:
        node_text = ws.cell(row, 4).value  # Node Text column
        type_col = ws.cell(row, 5).value   # Type column
        level = ws.cell(row, 3).value      # Level column
        unique_id = ws.cell(row, 6).value  # Unique ID column
        parent = ws.cell(row, 8).value     # Parent Node
        print(f'  Row {row}: Level {level} | Type: {type_col} | Text: "{node_text}" | Parent: "{parent}"')

print()
print('=== RELATIONSHIP: Type vs N/A in XPath ===')

# Analyze which node TYPES have N/A XPath
type_col = 5
xpath_col_num = xpath_col

type_na_map = {}
for row in range(2, max_row + 1):
    node_type = str(ws.cell(row, type_col).value)
    xpath_val = ws.cell(row, xpath_col_num).value
    
    if node_type not in type_na_map:
        type_na_map[node_type] = {'total': 0, 'with_na_xpath': 0}
    
    type_na_map[node_type]['total'] += 1
    if xpath_val == "N/A":
        type_na_map[node_type]['with_na_xpath'] += 1

print('Distribution of N/A XPath by Node Type:')
for node_type in sorted(type_na_map.keys()):
    stats = type_na_map[node_type]
    pct = (stats['with_na_xpath'] / stats['total'] * 100) if stats['total'] > 0 else 0
    print(f'  {node_type}: {stats["with_na_xpath"]}/{stats["total"]} have N/A XPath ({pct:.1f}%)')

print()
print('=== UNIQUE NODE TEXTS WITH N/A XPATH ===')
xpath_na_texts = {}
for row in xpath_na_rows:
    node_text = str(ws.cell(row, 4).value)
    node_type = str(ws.cell(row, 5).value)
    level = ws.cell(row, 3).value
    
    if node_text not in xpath_na_texts:
        xpath_na_texts[node_text] = {'count': 0, 'types': set(), 'levels': set()}
    
    xpath_na_texts[node_text]['count'] += 1
    xpath_na_texts[node_text]['types'].add(node_type)
    xpath_na_texts[node_text]['levels'].add(level)

print(f'Total unique node texts with N/A XPath: {len(xpath_na_texts)}')
print()
print('All node texts with N/A XPath:')
for text in sorted(xpath_na_texts.keys()):
    info = xpath_na_texts[text]
    types_str = ', '.join(sorted(info['types']))
    levels_str = ', '.join(str(l) for l in sorted(info['levels']))
    print(f'  "{text}"')
    print(f'    Count: {info["count"]} | Types: {types_str} | Levels: {levels_str}')

print()
print('=== PATTERN SUMMARY ===')
print(f'Total rows in file: {max_row - 1}')
print(f'Rows with N/A in XPath: {len(xpath_na_rows)} ({len(xpath_na_rows)/(max_row-1)*100:.2f}%)')
print()

# Check if N/A XPath nodes are ALL Parent (Expandable)
all_parent = all(str(ws.cell(row, 5).value) == "Parent (Expandable)" for row in xpath_na_rows)
print(f'Are ALL N/A XPath nodes "Parent (Expandable)"? {all_parent}')

# Check what types appear with N/A
types_with_na = set()
for row in xpath_na_rows:
    types_with_na.add(str(ws.cell(row, 5).value))
print(f'Node types with N/A XPath: {sorted(types_with_na)}')

# Check Unique ID pattern
print()
print('=== UNIQUE ID PATTERNS FOR N/A XPATH ROWS ===')
unique_ids_na = defaultdict(int)
for row in xpath_na_rows:
    unique_id = str(ws.cell(row, 6).value)
    unique_ids_na[unique_id] += 1

print(f'Unique ID patterns:')
for uid_pattern in sorted(unique_ids_na.keys())[:10]:
    print(f'  "{uid_pattern}": {unique_ids_na[uid_pattern]} occurrences')

print()
print('=== LOOKING FOR SPECIAL CHARACTERS IN IDS ===')
for uid_pattern in sorted(unique_ids_na.keys()):
    if 'NODE:' in uid_pattern:
        print(f'  Found NODE: prefix - "{uid_pattern}"')
