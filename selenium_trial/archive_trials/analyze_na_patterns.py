from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('menu_hierarchy3.xlsx')
ws = wb.active

print('=== FILE STRUCTURE ===')
print(f'Sheet name: {ws.title}')

# Get dimensions
max_row = ws.max_row
max_col = ws.max_column
print(f'Dimensions: {max_row} rows, {max_col} columns')
print()

# Print header
print('=== COLUMN HEADERS ===')
header = []
for col in range(1, max_col + 1):
    cell_value = ws.cell(1, col).value
    header.append(str(cell_value) if cell_value else 'Empty')
for i, h in enumerate(header):
    print(f'  Col {i+1}: {h}')
print()

# Print first 35 rows with NA analysis
print('=== FIRST 35 DATA ROWS (NA indicators shown) ===')
for row in range(2, min(37, max_row + 1)):
    values = []
    na_cols = []
    for col in range(1, max_col + 1):
        cell_value = ws.cell(row, col).value
        if cell_value is None:
            values.append('[NA]')
            na_cols.append(col)
        else:
            val_str = str(cell_value)[:35]
            values.append(val_str)
    
    # Show row details
    if na_cols:
        na_names = [header[c-1] for c in na_cols]
        print(f'\nRow {row}: >>> NAs in [{", ".join(na_names)}]')
    else:
        print(f'\nRow {row}: [All columns have values]')
    
    for col in range(1, max_col + 1):
        print(f'  {header[col-1]}: {values[col-1]}')

# Count NAs by column
print('\n\n=== NA ANALYSIS BY COLUMN ===')
total_data_rows = max_row - 1
for col in range(1, max_col + 1):
    col_name = header[col-1]
    na_count = 0
    for row in range(2, max_row + 1):
        if ws.cell(row, col).value is None:
            na_count += 1
    pct = (na_count / total_data_rows * 100) if total_data_rows > 0 else 0
    if na_count > 0:
        print(f'{col_name}: {na_count}/{total_data_rows} NAs ({pct:.1f}%)')
    else:
        print(f'{col_name}: 0 NAs (0.0%)')

# Check for patterns - which rows have NAs
print('\n\n=== ROWS WITH NA VALUES (complete list) ===')
rows_with_na = []
for row in range(2, max_row + 1):
    has_na = False
    for col in range(1, max_col + 1):
        if ws.cell(row, col).value is None:
            has_na = True
            break
    if has_na:
        rows_with_na.append(row)

print(f'Total rows with at least one NA: {len(rows_with_na)} out of {total_data_rows}')
print(f'Rows: {rows_with_na[:50]}')  # Show first 50

# Analyze what columns have NAs and their relationship
print('\n\n=== NA PATTERN ANALYSIS ===')
na_patterns = {}
for row in range(2, max_row + 1):
    na_cols = []
    for col in range(1, max_col + 1):
        if ws.cell(row, col).value is None:
            na_cols.append(header[col-1])
    
    if na_cols:
        pattern_key = tuple(sorted(na_cols))
        if pattern_key not in na_patterns:
            na_patterns[pattern_key] = []
        na_patterns[pattern_key].append(row)

print('Different NA patterns found:')
for pattern, rows in sorted(na_patterns.items(), key=lambda x: len(x[1]), reverse=True):
    print(f'\n  Pattern: {list(pattern)}')
    print(f'  Occurs in {len(rows)} rows')
    print(f'  Example rows: {rows[:5]}')

# Try to understand depth/level relationships
print('\n\n=== LOOKING AT DEPTH/HIERARCHY PATTERNS ===')
if 'Depth' in header or 'Level' in header:
    depth_col = None
    for col in range(1, max_col + 1):
        if header[col-1] in ['Depth', 'Level']:
            depth_col = col
            break
    
    if depth_col:
        print(f'Analyzing by {header[depth_col-1]} column:')
        depth_na_map = {}
        for row in range(2, max_row + 1):
            depth = ws.cell(row, depth_col).value
            has_na = any(ws.cell(row, c).value is None for c in range(1, max_col + 1))
            if depth not in depth_na_map:
                depth_na_map[depth] = {'total': 0, 'with_na': 0}
            depth_na_map[depth]['total'] += 1
            if has_na:
                depth_na_map[depth]['with_na'] += 1
        
        for depth in sorted(depth_na_map.keys()):
            stats = depth_na_map[depth]
            pct = (stats['with_na'] / stats['total'] * 100) if stats['total'] > 0 else 0
            print(f'  Depth/Level {depth}: {stats["with_na"]}/{stats["total"]} rows with NAs ({pct:.1f}%)')
