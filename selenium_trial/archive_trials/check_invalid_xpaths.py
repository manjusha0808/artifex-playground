import openpyxl

wb = openpyxl.load_workbook('uiMap_selenium_fullrun_auto3_cleaned.xlsx')
ws = wb.active

invalid = []
for i in range(2, ws.max_row + 1):
    xpath = ws[i][2].value
    if xpath and not str(xpath).startswith('//'):
        invalid.append((i, ws[i][0].value, xpath))

print(f'Found {len(invalid)} invalid XPaths (not starting with //):')
print()
for row, page, xpath in invalid[:15]:
    print(f'Row {row}: Page={page}')
    print(f'  XPath: {xpath}')
    print()
